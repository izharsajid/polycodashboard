import json, re, datetime
from openpyxl import load_workbook

SRC='/mnt/user-data/uploads/Polyco_Statement_Polyco_Version_12_feb_2026__1_.xlsx'
wb=load_workbook(SRC, data_only=True); ws=wb.active

def norm_date(v):
    """Return (iso, raw, flag) — flag set when the stored date is implausible."""
    if v is None: return None, None, None
    if isinstance(v, datetime.datetime):
        iso=v.date().isoformat()
        # statement as-at 2026-07-28; anything later is a day/month transposition
        if v.date() > datetime.date(2026,7,28):
            swapped=None
            try: swapped=datetime.date(v.year, v.day, v.month).isoformat()
            except ValueError: pass
            return swapped or iso, iso, "date-after-statement-date; day/month transposed in source"
        return iso, iso, None
    s=str(v).strip()
    m=re.match(r'^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$', s)
    if m:
        d,mo,y=int(m.group(1)),int(m.group(2)),int(m.group(3))
        y = y+2000 if y<100 else y
        try: return datetime.date(y,mo,d).isoformat(), s, None
        except ValueError: return None, s, "unparseable date"
    return None, s, "unparseable date"

def num(v):
    """Coerce a cell to a number. Placeholders such as '-' become None."""
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    t=str(v).strip().replace(',','').replace('$','')
    if t in ('','-','n/a','N/A'): return None
    try: return float(t)
    except ValueError: return None

def rtype(b, g, h):
    t=(b or "").strip().lower()
    if t.startswith("funds rcvd") or t.startswith("overpayment"): return "receipt"
    if "cargo clear" in t or "freight" in t or "clearnce" in t or "clearance" in t: return "recharge"
    if g not in (None,0): return "delivery"
    return "pending_po"

rows=[]
for r in range(7,181):
    b=ws.cell(r,2).value
    if b is None: continue
    d=num(ws.cell(r,4).value); f=num(ws.cell(r,6).value); g=num(ws.cell(r,7).value); h=num(ws.cell(r,8).value)
    di,draw,dflag = norm_date(ws.cell(r,9).value)
    mi,mraw,mflag = norm_date(ws.cell(r,13).value)
    po=re.search(r'2[0-9]{6}(?:-\d+)?', str(b))
    rows.append(dict(
        source_row=r, ref=str(b).strip(), po_number=po.group() if po else None,
        product=(str(ws.cell(r,3).value).strip() if ws.cell(r,3).value else None),
        type=rtype(b,g,h),
        po_amount=d, proforma_ref=(str(ws.cell(r,5).value).strip() if ws.cell(r,5).value else None),
        proforma_amount=f, delivered_value=g, received=h,
        received_date=di, received_date_source=draw,
        loaded=(str(ws.cell(r,10).value).strip() if ws.cell(r,10).value else None),
        delivery_date=mi, delivery_date_source=mraw,
        flags=[x for x in (dflag,mflag) if x]))

T=lambda k: round(sum(r[k] or 0 for r in rows),2)
summary=dict(
    as_at="2026-07-28",
    total_po_value=T("po_amount"), total_delivered=T("delivered_value"), total_received=T("received"),
    pos_pending_to_deliver=496489.00,
    containers_ready_next_month=137151.00, containers_in_process_following_month=69446.40)
summary["uncovered_advance"]=round(
    summary["total_received"]-summary["total_delivered"]
    -summary["pos_pending_to_deliver"]-summary["containers_ready_next_month"]
    -summary["containers_in_process_following_month"],2)
summary["recharges_included_in_delivered"]=round(sum(r["delivered_value"] or 0 for r in rows if r["type"]=="recharge"),2)

print(json.dumps(summary,indent=2))
print("rows",len(rows),"flagged dates",sum(1 for r in rows if r["flags"]))
assert abs(summary["uncovered_advance"]-1410206.34)<0.01, summary["uncovered_advance"]

json.dump(dict(source="EcoFibre x Polyco Ledger, statement workbook",currency="USD",
               summary=summary,rows=rows),
          open('/home/claude/ecofibre-polyco/data/polyco-ledger.json','w'),indent=2)
